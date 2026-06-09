import random
import unicodedata

import openpyxl
from datetime import datetime, date, timedelta


LINE_GROSS_CAPACITY_MIN = 480
END_OF_DAY_CLEANING_TIME_MIN = 30
CLEANING_OPERATORS = 5

LINE_AVAILABLE_CAPACITY_MIN = (
    LINE_GROSS_CAPACITY_MIN - END_OF_DAY_CLEANING_TIME_MIN
)


def calculate_production_time(n_boxes, cakes_per_box, cakes_per_hour_rate):
    if cakes_per_hour_rate is None or cakes_per_hour_rate == 0:
        return None

    n_cakes = n_boxes * cakes_per_box
    time_hours = n_cakes / cakes_per_hour_rate
    return time_hours * 60


def _safe_int(value, default=None):
    if value is None:
        return default

    if isinstance(value, str) and value.strip() in ["", "-"]:
        return default

    try:
        return int(value)
    except (ValueError, TypeError):
        return default


def _safe_float(value, default=None):
    if value is None:
        return default

    if isinstance(value, str) and value.strip() in ["", "-"]:
        return default

    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def _safe_text(value, default=None):
    if value is None:
        return default

    if isinstance(value, str) and value.strip() == "":
        return default

    return str(value).strip()


def _safe_time(value, default=None):
    if value is None:
        return default

    return value


def _normalize_label(value):
    if value is None:
        return ""

    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(
        char for char in text
        if not unicodedata.combining(char)
    )

    text = text.replace("\u00ba", "o")
    text = text.replace("\u00aa", "a")

    for char in ["_", "-", "/", "\\", "(", ")", "[", "]", ":", ";"]:
        text = text.replace(char, " ")

    return " ".join(text.split())


def _safe_date(value, default=None):
    if value is None:
        return default

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, str):
        value = value.strip()

        if value == "":
            return default

        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                pass

    return default


def _build_working_days(start_date, end_date, holidays=None):
    holidays = holidays or set()
    working_days = []
    current = start_date

    while current <= end_date:
        is_weekend = current.weekday() >= 5

        if not is_weekend and current not in holidays:
            working_days.append(current)

        current += timedelta(days=1)

    return working_days


def _read_holidays_sheet(ws):
    holidays = set()

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        holiday = _safe_date(row[0])

        if holiday is not None:
            holidays.add(holiday)

    return holidays


def _delivery_day_from_calendar(delivery_date, working_days):
    if delivery_date is None:
        return 1, None

    if not working_days:
        return 1, None

    adjusted_date = delivery_date

    while adjusted_date not in working_days:
        adjusted_date -= timedelta(days=1)

        if adjusted_date < working_days[0]:
            return 1, adjusted_date

    return working_days.index(adjusted_date) + 1, adjusted_date


def _is_yes(value):
    if value is None:
        return False

    val = str(value).strip().lower()
    return val == "yes" or val == "sim"


def _positive_value(value):
    return isinstance(value, (int, float)) and value > 0



def _find_header_indexes_generic(ws, required, aliases, max_scan_rows=30):
    normalized_aliases = {
        field: [_normalize_label(name) for name in names]
        for field, names in aliases.items()
    }

    for row_number, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_scan_rows), values_only=True),
        start=1
    ):
        labels = [_normalize_label(cell) for cell in row]
        indexes = {}

        for field, names in normalized_aliases.items():
            for index, label in enumerate(labels):
                if label in names:
                    indexes[field] = index
                    break

        if all(field in indexes for field in required):
            return row_number, indexes

    return None, {}



def _read_references_sheet(ws):
    refs = []
    incomplete_refs = []

    header_row, indexes = _find_header_indexes_generic(
        ws,
        required=["ref_id"],
        aliases={
            "ref_id": ["ref_id", "ref id", "referencia", "reference"],
            "name": ["nome", "name", "produto", "product"],
            "cakes_per_box": ["unid_caixa", "unid caixa", "cakes per box"],
            "family": ["familia", "família", "family"],
            "lead_time_L0_days": ["lead_time_l0_dias", "lead time l0 dias"],
            "can_L1": ["pode_l1", "pode l1", "can l1"],
            "rate_L1_prod": ["tempo_l1_prod_min cadencia de l1", "tempo l1 prod min cadencia de l1", "rate l1 prod"],
            "ops_L1_finish": ["operadores_nec_l1_acab", "operadores nec l1 acab"],
            "ops_L1_prod": ["operadores_nec_l1_prod", "operadores nec l1 prod"],
            "can_L2": ["pode_l2", "pode l2", "can l2"],
            "rate_L2_prod": ["tempo_l2_prod_min cadencia de l2", "tempo l2 prod min cadencia de l2", "rate l2 prod"],
            "rate_L2_finish": ["tempo_l2_acab_min cadencia de l2", "tempo l2 acab min cadencia de l2", "rate l2 finish"],
            "ops_L2_finish": ["operadores_nec_l2_acab", "operadores nec l2 acab"],
            "ops_L2_prod": ["operadores_nec_l2_prod", "operadores nec l2 prod"],
            "kg_per_master_box": ["kg", "kg por caixa", "kg per master box"],
            "economic_value_per_master_box": ["euros", "valor economico", "economic value", "economic value per master box"],
            "needs_oven": ["precisa_forno", "precisa forno", "requires oven"],
            "monday_forbidden": [
                "nao_produzir_segunda",
                "nao produzir segunda",
                "nao pode segunda",
                "monday forbidden",
                "cannot monday",
            ],
        }
    )

    if header_row is None:
        raise ValueError("Could not find ref_id header in sheet 2_REFERENCIAS.")

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        ref_id = _get_row_value(row, indexes, "ref_id")

        if ref_id is None:
            continue

        ref_id = str(ref_id).strip()
        name = _safe_text(_get_row_value(row, indexes, "name"), default=ref_id)
        cakes_per_box = _safe_int(_get_row_value(row, indexes, "cakes_per_box"), default=1)
        family = _safe_text(_get_row_value(row, indexes, "family"), default="no_family").lower()
        kg_per_master_box = _safe_float(_get_row_value(row, indexes, "kg_per_master_box"), default=0)
        economic_value_per_master_box = _safe_float(_get_row_value(row, indexes, "economic_value_per_master_box"), default=0)
        needs_oven = _is_yes(_get_row_value(row, indexes, "needs_oven"))
        monday_forbidden = _safe_int(
            _get_row_value(row, indexes, "monday_forbidden"),
            default=0
        ) == 1

        can_L1 = _is_yes(_get_row_value(row, indexes, "can_L1"))
        rate_L1_prod = _safe_float(_get_row_value(row, indexes, "rate_L1_prod"))
        ops_L1_finish = _safe_int(_get_row_value(row, indexes, "ops_L1_finish"))
        ops_L1_prod = _safe_int(_get_row_value(row, indexes, "ops_L1_prod"))
        if can_L1:
            if rate_L1_prod is None:
                incomplete_refs.append((ref_id, "can_L1=Yes but no L1 rate"))
            if ops_L1_prod is None:
                incomplete_refs.append((ref_id, "can_L1=Yes but no production operators L1"))
            if ops_L1_finish is None:
                incomplete_refs.append((ref_id, "can_L1=Yes but no finishing operators L1"))
        else:
            rate_L1_prod = 0
            ops_L1_prod = 0
            ops_L1_finish = 0

        can_L2 = _is_yes(_get_row_value(row, indexes, "can_L2"))
        rate_L2_prod = _safe_float(_get_row_value(row, indexes, "rate_L2_prod"))
        rate_L2_finish = _safe_float(_get_row_value(row, indexes, "rate_L2_finish"))
        ops_L2_finish = _safe_int(_get_row_value(row, indexes, "ops_L2_finish"))
        ops_L2_prod = _safe_int(_get_row_value(row, indexes, "ops_L2_prod"))

        if can_L2:
            if rate_L2_prod is None:
                incomplete_refs.append((ref_id, "can_L2=Yes but no production rate L2"))
            if rate_L2_finish is None:
                incomplete_refs.append((ref_id, "can_L2=Yes but no finishing rate L2"))
            if ops_L2_prod is None:
                incomplete_refs.append((ref_id, "can_L2=Yes but no production operators L2"))
            if ops_L2_finish is None:
                incomplete_refs.append((ref_id, "can_L2=Yes but no finishing operators L2"))
        else:
            rate_L2_prod = 0
            rate_L2_finish = 0
            ops_L2_prod = 0
            ops_L2_finish = 0

        fixed_line = None
        if can_L1 and _positive_value(rate_L1_prod):
            fixed_line = "L1"
        elif can_L2 and _positive_value(rate_L2_prod):
            fixed_line = "L2"

        if can_L1 and can_L2:
            incomplete_refs.append((ref_id, "both L1 and L2 are available; fixed_line set to L1"))

        refs.append({
            "id": ref_id,
            "name": name,
            "family": family,
            "cakes_per_box": cakes_per_box,
            "kg_per_master_box": kg_per_master_box,
            "economic_value_per_master_box": economic_value_per_master_box,
            "needs_oven": needs_oven,
            "monday_forbidden": monday_forbidden,
            "fixed_line": fixed_line,
            "lead_time_L0_days": _safe_int(_get_row_value(row, indexes, "lead_time_L0_days"), default=1),
            "can_L1": can_L1,
            "rate_L1_prod": rate_L1_prod,
            "rate_L1_finish": rate_L1_prod,
            "ops_L1_prod": ops_L1_prod,
            "ops_L1_finish": ops_L1_finish,
            "can_L2": can_L2,
            "rate_L2_prod": rate_L2_prod,
            "rate_L2_finish": rate_L2_finish,
            "ops_L2_prod": ops_L2_prod,
            "ops_L2_finish": ops_L2_finish,
        })

    return refs, incomplete_refs


    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue

        ref_id = str(row[0]).strip()
        name = str(row[1]).strip() if row[1] else ref_id
        cakes_per_box = _safe_int(row[2], default=1)
        family = str(row[3]).strip().lower() if row[3] else "no_family"

        can_L1 = _is_yes(row[9])
        rate_L1_prod = _safe_float(row[10])
        ops_L1_finish = _safe_int(row[11])
        ops_L1_prod = _safe_int(row[12])

        if can_L1:
            if rate_L1_prod is None:
                incomplete_refs.append((ref_id, "can_L1=Yes but no L1 rate"))
            if ops_L1_prod is None:
                incomplete_refs.append((ref_id, "can_L1=Yes but no production operators L1"))
            if ops_L1_finish is None:
                incomplete_refs.append((ref_id, "can_L1=Yes but no finishing operators L1"))
        else:
            rate_L1_prod = 0
            ops_L1_prod = 0
            ops_L1_finish = 0

        can_L2 = _is_yes(row[13])
        rate_L2_prod = _safe_float(row[14])
        rate_L2_finish = _safe_float(row[15])
        ops_L2_finish = _safe_int(row[16])
        ops_L2_prod = _safe_int(row[17])

        if can_L2:
            if rate_L2_prod is None:
                incomplete_refs.append((ref_id, "can_L2=Yes but no production rate L2"))
            if rate_L2_finish is None:
                incomplete_refs.append((ref_id, "can_L2=Yes but no finishing rate L2"))
            if ops_L2_prod is None:
                incomplete_refs.append((ref_id, "can_L2=Yes but no production operators L2"))
            if ops_L2_finish is None:
                incomplete_refs.append((ref_id, "can_L2=Yes but no finishing operators L2"))
        else:
            rate_L2_prod = 0
            rate_L2_finish = 0
            ops_L2_prod = 0
            ops_L2_finish = 0

        ref = {
            "id": ref_id,
            "name": name,
            "family": family,
            "cakes_per_box": cakes_per_box,
            "lead_time_L0_days": _safe_int(row[8], default=1),

            "can_L1": can_L1,
            "rate_L1_prod": rate_L1_prod,
            "rate_L1_finish": rate_L1_prod,
            "ops_L1_prod": ops_L1_prod,
            "ops_L1_finish": ops_L1_finish,

            "can_L2": can_L2,
            "rate_L2_prod": rate_L2_prod,
            "rate_L2_finish": rate_L2_finish,
            "ops_L2_prod": ops_L2_prod,
            "ops_L2_finish": ops_L2_finish,
        }

        refs.append(ref)

    return refs, incomplete_refs


def _read_structure_sheet(ws):
    structure = {
        "n_days": None,
        "planning_start_date": None,
        "planning_end_date": None,
        "line_capacity_min": LINE_GROSS_CAPACITY_MIN,
        "end_of_day_cleaning_time_min": END_OF_DAY_CLEANING_TIME_MIN,
        "cleaning_operators": CLEANING_OPERATORS,
        "capacity_L0_min": None,
        "lead_time_standard_L0_L1L2_days": None,
        "start_time_L0": None,
        "end_time_L0": None,
        "n_ovens": None,
        "ovens_capacity_min": None,
        "start_time_L1_prod": None,
        "end_time_L1_prod": None,
        "capacity_L1_prod_min": None,
        "start_time_L1_finish": None,
        "end_time_L1_finish": None,
        "capacity_L1_finish_min": None,
        "tunnel_time_L1_min": None,
        "start_time_L2_prod": None,
        "end_time_L2_prod": None,
        "capacity_L2_prod_min": None,
        "start_time_L2_finish": None,
        "end_time_L2_finish": None,
        "capacity_L2_finish_min": None,
        "nitrogen_time_L2_min": None,
        "n_productive_operators": None,
        "n_operators": None,
        "operators_rotate_L0_L1_L2": None,
        "_extra": {},
    }

    param_to_key = {
        "number of working days": ("n_days", "int"),
        "working days": ("n_days", "int"),
        "1º dia do planeamento": ("planning_start_date", "date"),
        "último dia a planear": ("planning_end_date", "date"),
        "effective capacity of l0 per day (minutes)": ("capacity_L0_min", "int"),
        "standard lead time l0 → l1/l2 (days)": ("lead_time_standard_L0_L1L2_days", "int"),
        "standard lead time l0 â†’ l1/l2 (days)": ("lead_time_standard_L0_L1L2_days", "int"),
        "start time of l0": ("start_time_L0", "time"),
        "end time of l0": ("end_time_L0", "time"),
        "number of available ovens": ("n_ovens", "int"),
        "effective capacity per oven per day (minutes)": ("ovens_capacity_min", "int"),
        "start time l1 production": ("start_time_L1_prod", "time"),
        "end time l1 production": ("end_time_L1_prod", "time"),
        "effective capacity l1 production (minutes)": ("capacity_L1_prod_min", "int"),
        "start time l1 finishing/packaging": ("start_time_L1_finish", "time"),
        "end time l1 finishing/packaging": ("end_time_L1_finish", "time"),
        "effective capacity l1 finishing/packaging (minutes)": ("capacity_L1_finish_min", "int"),
        "time of cooling tunnel l1 (minutes)": ("tunnel_time_L1_min", "int"),
        "start time l2 production": ("start_time_L2_prod", "time"),
        "end time l2 production": ("end_time_L2_prod", "time"),
        "effective capacity l2 production (minutes)": ("capacity_L2_prod_min", "int"),
        "start time l2 finishing/packaging": ("start_time_L2_finish", "time"),
        "end time l2 finishing/packaging": ("end_ time_L2_finish", "time"),
        "effective capacity l2 finishing/packaging (minutes)": ("capacity_L2_finish_min", "int"),
        "time of nitrogen chamber l2 (minutes)": ("nitrogen_time_L2_min", "int"),
        "total number of productive operators": ("n_productive_operators", "int"),
        "numero total de operadores produtivos": ("n_productive_operators", "int"),
        "do operators rotate between l0/l1/l2?": ("operators_rotate_L0_L1_L2", "bool"),
    }

    extra_param_to_key = {
        "1\u00ba dia do planeamento": ("planning_start_date", "date"),
        "ultimo dia a planear": ("planning_end_date", "date"),
        "numero total de operadores produtivos": ("n_productive_operators", "int"),
        "capacidade efetiva de l0 por dia minutos": ("capacity_L0_min", "int"),
        "numero de fornos disponiveis": ("n_ovens", "int"),
    }

    param_to_key = {
        _normalize_label(key): value
        for key, value in param_to_key.items()
    }

    param_to_key.update({
        _normalize_label(key): value
        for key, value in extra_param_to_key.items()
    })

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue

        param = _normalize_label(row[0])
        value = row[1]

        mapping = param_to_key.get(param)

        if mapping is not None:
            key, type_ = mapping

            if type_ == "int":
                structure[key] = _safe_int(value, default=structure[key])
            elif type_ == "bool":
                structure[key] = _is_yes(value)
            elif type_ == "time":
                structure[key] = _safe_time(value, default=structure[key])
            elif type_ == "date":
                structure[key] = _safe_date(value, default=structure[key])
            else:
                structure[key] = _safe_text(value, default=structure[key])

        elif value is not None:
            structure["_extra"][param] = value

    return structure


def _read_operators_sheet(ws):
    operators = []

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue

        in_pool = _is_yes(row[2])

        if not in_pool:
            continue

        availability = [
            _safe_int(row[3 + i], default=0)
            for i in range(5)
            if 3 + i < len(row)
        ]

        operators.append({
            "id": str(row[0]).strip(),
            "name": str(row[1]).strip() if row[1] else str(row[0]).strip(),
            "availability": availability,
        })

    return operators


def _read_family_aliases_sheet(ws):
    aliases = {}

    header_row, indexes = _find_header_indexes_generic(
        ws,
        required=["family"],
        aliases={
            "family": ["familia", "família", "family"],
        }
    )

    if header_row is None:
        return aliases

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        family = _safe_text(_get_row_value(row, indexes, "family"))

        if family:
            family = family.lower()
            aliases[_normalize_label(family)] = family

    return aliases


    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        family = _safe_text(row[3] if len(row) > 3 else None)

        if family:
            family = family.lower()
            aliases[_normalize_label(family)] = family

        family_name = _safe_text(row[20] if len(row) > 20 else None)
        family_code = _safe_text(row[21] if len(row) > 21 else None)

        if family_name:
            family_name = family_name.lower()
            aliases[_normalize_label(family_name)] = family_name

            if family_code:
                aliases[_normalize_label(family_code)] = family_name

    return aliases


def _canonical_family(value, family_aliases):
    if value is None:
        return None

    text = str(value).strip().lower()
    key = _normalize_label(text)

    return family_aliases.get(key, text)



def _read_setups_sheet(ws, all_families, family_aliases=None):
    matrix = {}
    family_aliases = family_aliases or {}
    valid_families = set(all_families)

    for family in all_families:
        family_aliases[_normalize_label(family)] = family

    header = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]

    column_families = [
        _canonical_family(c, family_aliases)
        for c in header[1:]
    ]

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue

        from_family = _canonical_family(row[0], family_aliases)

        for j, to_family in enumerate(column_families):
            if from_family is None or to_family is None:
                continue

            if from_family not in valid_families or to_family not in valid_families:
                continue

            value = row[j + 1] if j + 1 < len(row) else None

            if isinstance(value, (int, float)):
                matrix[(from_family, to_family)] = float(value)

    return matrix, 0


def _find_header_indexes(ws):
    aliases = {
        "ref_id": ["ref id", "referencia", "reference", "produto"],
        "master_boxes": [
            "master boxes",
            "caixas master",
            "quantidade",
            "forecast unid",
            "forecast unidades",
        ],
        "delivery_date": ["data entrega", "delivery date", "due date"],
        "priority": ["prioridade", "priority"],
    }

    for row_number, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True),
        start=1
    ):
        labels = [_normalize_label(cell) for cell in row]
        indexes = {}

        for field, names in aliases.items():
            normalized_names = [_normalize_label(name) for name in names]

            for index, label in enumerate(labels):
                if label in normalized_names:
                    indexes[field] = index
                    break

        if "ref_id" in indexes:
            return row_number, indexes

    return None, {}


def _get_row_value(row, indexes, field, default_index=None):
    index = indexes.get(field, default_index)

    if index is None or index >= len(row):
        return None

    return row[index]


def _read_demand_sheet(ws, working_days=None):
    demand = []
    header_row, indexes = _find_header_indexes(ws)

    if header_row is None:
        header_row = 4
        indexes = {
            "ref_id": 0,
            "master_boxes": 1,
            "delivery_date": 2,
            "priority": 3,
        }

    for row in ws.iter_rows(
        min_row=header_row + 1,
        max_row=ws.max_row,
        values_only=True
    ):
        ref_id = _get_row_value(row, indexes, "ref_id")

        if ref_id is None:
            continue

        delivery_value = _get_row_value(row, indexes, "delivery_date")
        delivery_calendar_date = _safe_date(delivery_value)

        has_calendar_horizon = (
            working_days
            and working_days[0] is not None
        )

        if has_calendar_horizon and delivery_calendar_date:
            delivery_date, adjusted_delivery_date = _delivery_day_from_calendar(
                delivery_calendar_date,
                working_days
            )
        else:
            delivery_date = _safe_int(delivery_value, default=1)
            adjusted_delivery_date = None

        order = {
            "ref_id": str(ref_id).strip(),
            "master_boxes": _safe_int(
                _get_row_value(row, indexes, "master_boxes"),
                default=0
            ),
            "delivery_calendar_date": delivery_calendar_date,
            "adjusted_delivery_date": adjusted_delivery_date,
            "delivery_date": delivery_date,
            "priority": _safe_text(
                _get_row_value(row, indexes, "priority"),
                default="Medium"
            ),
        }

        demand.append(order)

    return demand


def _generate_synthetic_demand(refs, n_days, n_orders=15, seed=42):
    random.seed(seed)

    valid_refs = [
        r for r in refs
        if (
            r["can_L1"] and _positive_value(r["rate_L1_prod"])
        ) or (
            r["can_L2"] and _positive_value(r["rate_L2_prod"])
        )
    ]

    if not valid_refs:
        print("No valid references found to generate synthetic demand.")
        return []

    demand = []

    for _ in range(n_orders):
        ref = random.choice(valid_refs)

        order = {
            "ref_id": ref["id"],
            "master_boxes": random.choice([100, 150, 200, 300, 500]),
            "delivery_calendar_date": None,
            "adjusted_delivery_date": None,
            "delivery_date": random.randint(2, n_days),
            "priority": random.choice(["High", "Medium", "Low"]),
        }

        demand.append(order)

    return demand


def _normalize_key(value):
    if value is None:
        return None

    text = str(value).strip().lower()

    replacements = {
        "á": "a", "à": "a", "ã": "a", "â": "a",
        "é": "e", "ê": "e",
        "í": "i",
        "ó": "o", "ô": "o", "õ": "o",
        "ú": "u",
        "ç": "c",
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    text = text.replace("/", "_")
    text = text.replace("-", "_")
    text = text.replace(" ", "_")

    while "__" in text:
        text = text.replace("__", "_")

    return text



def _read_machines_sheet(ws):
    machines = {}
    ref_machine_requirements = {}

    header = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    header_labels = [_normalize_label(cell) for cell in header]

    machine_name_col = 0
    quantity_col = 1
    section_col = 2

    for idx, label in enumerate(header_labels):
        if label == "nome":
            machine_name_col = idx
            break

    for idx, label in enumerate(header_labels):
        if label in ["quantid total", "quantidade total", "total quantity"]:
            quantity_col = idx
        elif label in ["seccao", "secao", "section"]:
            section_col = idx

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        machine_name = row[machine_name_col] if machine_name_col < len(row) else None
        quantity = _safe_int(row[quantity_col] if quantity_col < len(row) else None, default=None)
        section = _safe_text(row[section_col] if section_col < len(row) else None, default=None)

        if machine_name is None or quantity is None:
            continue

        machine_id = _normalize_key(machine_name)

        machines[machine_id] = {
            "name": str(machine_name).strip(),
            "quantity": quantity,
            "section": section,
        }

    ref_id_col = None

    for idx, label in enumerate(header_labels):
        if label in ["ref id", "ref_id"]:
            ref_id_col = idx
            break

    if ref_id_col is None:
        return machines, ref_machine_requirements

    machine_columns = []

    for col_idx in range(ref_id_col + 1, len(header)):
        machine_name = header[col_idx]

        if machine_name is None:
            continue

        machine_id = _normalize_key(machine_name)
        machine_columns.append((col_idx, machine_id))

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        ref_id = row[ref_id_col] if ref_id_col < len(row) else None

        if ref_id is None:
            continue

        ref_id = str(ref_id).strip()
        requirements = {}

        for col_idx, machine_id in machine_columns:
            quantity_needed = _safe_int(row[col_idx] if col_idx < len(row) else None, default=0)

            if quantity_needed > 0:
                requirements[machine_id] = quantity_needed

        ref_machine_requirements[ref_id] = requirements

    return machines, ref_machine_requirements


    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        machine_name = row[1]
        quantity = _safe_int(row[2], default=None)
        section = _safe_text(row[3], default=None)

        if machine_name is None or quantity is None:
            continue

        machine_id = _normalize_key(machine_name)

        machines[machine_id] = {
            "name": str(machine_name).strip(),
            "quantity": quantity,
            "section": section,
        }

    header = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]

    machine_columns = []

    for col_idx in range(7, 24):
        machine_name = header[col_idx]

        if machine_name is None:
            continue

        machine_id = _normalize_key(machine_name)

        machine_columns.append((col_idx, machine_id))

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        ref_id = row[6]

        if ref_id is None:
            continue

        ref_id = str(ref_id).strip()

        requirements = {}

        for col_idx, machine_id in machine_columns:
            quantity_needed = _safe_int(row[col_idx], default=0)

            if quantity_needed > 0:
                requirements[machine_id] = quantity_needed

        ref_machine_requirements[ref_id] = requirements

    return machines, ref_machine_requirements

def _split_large_orders(demand, refs, final_lines, available_line_time_min):
    refs_by_id = {
        str(ref["id"]).strip(): ref
        for ref in refs
    }

    split_demand = []

    for order in demand:
        ref_id = str(order["ref_id"]).strip()

        if ref_id not in refs_by_id:
            split_demand.append(order)
            continue

        ref = refs_by_id[ref_id]

        possible_rates = []

        if "L1" in final_lines and ref["can_L1"] and _positive_value(ref["rate_L1_prod"]):
            possible_rates.append(ref["rate_L1_prod"])

        if "L2" in final_lines and ref["can_L2"] and _positive_value(ref["rate_L2_prod"]):
            possible_rates.append(ref["rate_L2_prod"])

        if not possible_rates:
            split_demand.append(order)
            continue

        fastest_rate = max(possible_rates)

        max_cakes_per_lot = (
            available_line_time_min / 60
        ) * fastest_rate

        max_boxes_per_lot = int(
            max_cakes_per_lot / ref["cakes_per_box"]
        )

        if max_boxes_per_lot <= 0:
            split_demand.append(order)
            continue

        total_boxes = order["master_boxes"]

        if total_boxes <= max_boxes_per_lot:
            new_order = dict(order)
            new_order["original_ref_id"] = ref_id
            new_order["lot_number"] = 1
            new_order["n_lots"] = 1
            split_demand.append(new_order)
            continue

        remaining_boxes = total_boxes
        lot_number = 1

        n_lots = (
            total_boxes + max_boxes_per_lot - 1
        ) // max_boxes_per_lot

        while remaining_boxes > 0:
            lot_boxes = min(max_boxes_per_lot, remaining_boxes)

            new_order = dict(order)
            new_order["ref_id"] = ref_id
            new_order["master_boxes"] = lot_boxes
            new_order["original_ref_id"] = ref_id
            new_order["lot_number"] = lot_number
            new_order["n_lots"] = n_lots

            split_demand.append(new_order)

            remaining_boxes -= lot_boxes
            lot_number += 1

    return split_demand

def load_real_instance(
    excel_path="Inputs_EmpresaX.xlsx",
    n_synthetic_orders=15,
    seed=42
):
    print(f"Loading instance from {excel_path}...")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    structure = _read_structure_sheet(wb["1_ESTRUTURA"])

    holidays = set()

    holidays_sheet = next(
        (name for name in wb.sheetnames if "FERIAD" in _normalize_label(name).upper()),
        None
    )

    if holidays_sheet is not None:
        holidays = _read_holidays_sheet(wb[holidays_sheet])

    if not structure["planning_start_date"] or not structure["planning_end_date"]:
        raise ValueError(
            "Planning start date and planning end date must be defined in "
            "sheet 1_ESTRUTURA using '1º dia do planeamento' and "
            "'Último dia a planear'."
        )

    working_days = _build_working_days(
        structure["planning_start_date"],
        structure["planning_end_date"],
        holidays
    )

    structure["n_days"] = len(working_days)

    refs, incomplete_refs = _read_references_sheet(wb["2_REFERENCIAS"])
    operators = []
   

    families = sorted(set(r["family"] for r in refs))
    family_aliases = _read_family_aliases_sheet(wb["2_REFERENCIAS"])

    setups_matrix, n_setups_estimated = _read_setups_sheet(
        wb["3_SETUPS"],
        families,
        family_aliases=family_aliases
    )

    machines_sheet = next(
        (
            name for name in wb.sheetnames
            if "maquinas" in _normalize_label(name) or name.startswith("7_") or name.startswith("10_")
        ),
        None
    )

    if machines_sheet is not None:
        machines, ref_machine_requirements = _read_machines_sheet(
            wb[machines_sheet]
        )
        machines_source = f"Excel sheet {machines_sheet}"
    else:
        machines = {}
        ref_machine_requirements = {}
        machines_source = "not available"

    demand_sheet = next(
        (name for name in wb.sheetnames if "procura" in _normalize_label(name)),
        None
    )

    if demand_sheet is not None:
        demand = _read_demand_sheet(
            wb[demand_sheet],
            working_days=working_days
        )

        if demand:
            demand_source = f"Excel sheet {demand_sheet}"
        else:
            demand = _generate_synthetic_demand(
                refs,
                structure["n_days"],
                n_orders=n_synthetic_orders,
                seed=seed
            )
            demand_source = f"synthetic because {demand_sheet} is empty"
    else:
        demand = _generate_synthetic_demand(
            refs,
            structure["n_days"],
            n_orders=n_synthetic_orders,
            seed=seed
        )
        demand_source = "synthetic because demand sheet does not exist"

    standard_operators = structure["n_productive_operators"]

    if standard_operators is None:
        raise ValueError(
            "Numero total de operadores produtivos must be defined in sheet 1_ESTRUTURA."
        )

    demand_before_split = len(demand)

    demand = _split_large_orders(
        demand,
        refs,
        ["L1", "L2"],
        structure["line_capacity_min"] - structure["end_of_day_cleaning_time_min"]
    )

    demand_after_split = len(demand)

    instance = {
        "n_days": structure["n_days"],
        "final_lines": ["L1", "L2"],
        "days": [f"day_{i + 1}" for i in range(structure["n_days"])],
        "working_days": working_days,
        "holidays": sorted(holidays),
        "standard_operators": standard_operators,
        "line_capacity_min": structure["line_capacity_min"],
        "end_of_day_cleaning_time_min": structure["end_of_day_cleaning_time_min"],
        "cleaning_operators": structure["cleaning_operators"],

        "available_line_time_min": (
            structure["line_capacity_min"]
            - structure["end_of_day_cleaning_time_min"]
        ),
        "time_bucket_min": 30,
        "monday_days": [
            i + 1
            for i, working_day in enumerate(working_days)
            if working_day.weekday() == 0
        ],

        "refs": refs,
        "families": families,
        "setups_matrix": setups_matrix,
        "operators": operators,
        "demand": demand,
        "machines": machines,
        "ref_machine_requirements": ref_machine_requirements,
        "structure": structure,

        "_meta": {
            "n_refs_total": len(refs),
            "n_incomplete_refs": len(incomplete_refs),
            "incomplete_refs": incomplete_refs,
            "n_families": len(families),
            "n_setups_estimated": n_setups_estimated,
            "demand_source": demand_source,
            "machines_source": machines_source,
            "n_machines": len(machines),
            "n_refs_with_machine_requirements": len(ref_machine_requirements),
            "n_orders_before_split": demand_before_split,
            "n_orders_after_split": demand_after_split,

        }
    }

    return instance


def print_instance_summary(instance):
    meta = instance["_meta"]

    print("=" * 70)
    print("LOADED INSTANCE - SUMMARY")
    print("=" * 70)

    print("\nSTRUCTURE")
    print(f"  Horizon: {instance['n_days']} working days")

    if instance["working_days"] and instance["working_days"][0] is not None:
        print(f"  Planning start: {instance['working_days'][0]}")
        print(f"  Planning end: {instance['working_days'][-1]}")
        print(f"  Holidays/non-working days: {len(instance['holidays'])}")

    print(f"  Gross capacity Line 1/Line 2: {instance['line_capacity_min']} min/day")
    print(
        f"  End of day cleaning: {instance['end_of_day_cleaning_time_min']} min "
        f"({instance['cleaning_operators']} operators)"
    )
    print(
        f"  Available for production: "
        f"{instance['available_line_time_min']} min/day"
    )
    print(f"  Standard operators per day: {instance['standard_operators']}")

    print("\nREFERENCES")
    print(f"  Total: {meta['n_refs_total']}")

    n_can_L1 = sum(1 for r in instance["refs"] if r["can_L1"])
    n_can_L2 = sum(1 for r in instance["refs"] if r["can_L2"])
    n_only_L1 = sum(1 for r in instance["refs"] if r["can_L1"] and not r["can_L2"])
    n_only_L2 = sum(1 for r in instance["refs"] if r["can_L2"] and not r["can_L1"])
    n_both = sum(1 for r in instance["refs"] if r["can_L1"] and r["can_L2"])

    print(f"    Can L1: {n_can_L1} (only L1: {n_only_L1})")
    print(f"    Can L2: {n_can_L2} (only L2: {n_only_L2})")
    print(f"    Can both: {n_both}")

    if meta["n_incomplete_refs"] > 0:
        print(f"  Incomplete refs: {meta['n_incomplete_refs']}")

        for ref_id, reason in meta["incomplete_refs"][:15]:
            print(f"      - {ref_id}: {reason}")

        if meta["n_incomplete_refs"] > 15:
            print("      ...")

    print("\nFAMILIES")
    print(f"  Total: {meta['n_families']}")
    print(
        f"  List: {', '.join(instance['families'][:10])}"
        f"{'...' if len(instance['families']) > 10 else ''}"
    )

    print("\nSETUPS")
    n_total_setups = len(instance["setups_matrix"])
    n_real = n_total_setups - meta["n_setups_estimated"]

    print(
        f"  Matrix: {meta['n_families']} x {meta['n_families']} "
        f"= {n_total_setups} values"
    )
    print(f"  Filled in Excel: {n_real}")
    print(f"  Estimated/default values: {meta['n_setups_estimated']}")

    print("\nOPERATORS")
    print(f"  Operators in shared pool: {len(instance['operators'])}")

    if instance["operators"]:
        max_availability_days = min(
            instance["n_days"],
            min(len(op["availability"]) for op in instance["operators"])
        )

        available_per_day = [
            sum(op["availability"][d] for op in instance["operators"])
            for d in range(max_availability_days)
        ]

        day_names = [
            f"day_{i + 1}"
            for i in range(max_availability_days)
        ]

        print(f"  Availability from sheet: {dict(zip(day_names, available_per_day))}")

        if max_availability_days < instance["n_days"]:
            print(
                "  Note: operator availability sheet has fewer days than "
                "the planning horizon. The optimization model should use "
                "standard_operators or the Excel sheet should be extended."
            )

    print("\nMACHINES / CRITICAL RESOURCES")
    print(f"  Total machines/resources: {meta['n_machines']}")
    print(f"  Source: {meta['machines_source']}")
    print(
        f"  References with machine requirements: "
        f"{meta['n_refs_with_machine_requirements']}"
    )

    if instance["machines"]:
        print("  First machines:")

        for machine_id, data in list(instance["machines"].items())[:10]:
            print(
                f"    {machine_id}: "
                f"{data['quantity']} available "
                f"({data['section']})"
            )

    print("\nDEMAND")
    print(f"  Orders: {len(instance['demand'])} ({meta['demand_source']})")
    print(
        f"  Orders after lot splitting: "
        f"{meta['n_orders_after_split']} "
        f"(from {meta['n_orders_before_split']} demand rows)"
    )

    if instance["demand"]:
        print("  First 5 orders:")

        for o in instance["demand"][:5]:
            print(
                f"    {o['ref_id']}: {o['master_boxes']} master boxes, "
                f"delivery day {o['delivery_date']}, "
                f"calendar due {o.get('delivery_calendar_date')}, "
                f"adjusted due {o.get('adjusted_delivery_date')}, "
                f"priority {o['priority']}"
            )

    print()


if __name__ == "__main__":
    instance = load_real_instance(excel_path="../Inputs_EmpresaX_small.xlsx")
    print_instance_summary(instance)

