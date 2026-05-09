import random
import openpyxl
import os

CAPACIDADE_LINHA_BRUTA_MIN = 480
TEMPO_LIMPEZA_FIM_DIA_MIN = 30
OPERADORES_LIMPEZA = 5

CAPACIDADE_LINHA_DISPONIVEL_MIN = (
    CAPACIDADE_LINHA_BRUTA_MIN - TEMPO_LIMPEZA_FIM_DIA_MIN
)

def calcular_tempo_producao(n_caixas, bolos_por_caixa, cadencia_bolos_h):
   
    if cadencia_bolos_h is None or cadencia_bolos_h == 0:
        return None

    n_bolos = n_caixas * bolos_por_caixa
    tempo_horas = n_bolos / cadencia_bolos_h
    return tempo_horas * 60


def _safe_int(valor, default=None):
    """
    Converte valor em int.
    Se estiver vazio ou '-', devolve default.
    """
    if valor is None:
        return default

    if isinstance(valor, str) and valor.strip() in ["", "-"]:
        return default

    try:
        return int(valor)
    except (ValueError, TypeError):
        return default


def _safe_float(valor, default=None):
    """
    Converte valor em float.
    Se estiver vazio ou '-', devolve default.
    """
    if valor is None:
        return default

    if isinstance(valor, str) and valor.strip() in ["", "-"]:
        return default

    try:
        return float(valor)
    except (ValueError, TypeError):
        return default


def _safe_text(valor, default=None):
    """
    Converte valor em string limpa.
    Mantém objetos de tempo como estão.
    """
    if valor is None:
        return default

    if isinstance(valor, str) and valor.strip() == "":
        return default

    return str(valor).strip()


def _safe_time(valor, default=None):
    """
    Retorna valores de tempo ou texto de horário.
    """
    if valor is None:
        return default

    return valor


def _eh_sim(valor):
    """
    Interpreta valores do Excel como Sim/Não.
    Aceita 'Sim', 'sim', 'SIM', etc.
    """
    if valor is None:
        return False

    return str(valor).strip().lower() == "sim"


def _valor_positivo(valor):
    """
    Verifica se um valor é numérico e maior do que zero.
    Evita erro quando a cadência é None.
    """
    return isinstance(valor, (int, float)) and valor > 0


# ============================================================
# LEITURA DA ABA 2_REFERENCIAS
# ============================================================

def _ler_aba_referencias(ws):
    """
    Lê a aba 2_REFERENCIAS.

    Estrutura esperada:
        Col 0: ref_id
        Col 1: nome
        Col 2: Unid_caixa
        Col 3: família
        Col 9: lead_time_L0_dias
        Col 10: pode_L1
        Col 11: cadencia_L1_prod
        Col 12: operadores_nec_L1_acab
        Col 13: operadores_nec_L1_prod
        Col 14: pode_L2
        Col 15: cadencia_L2_prod
        Col 16: cadencia_L2_acab
        Col 17: operadores_nec_L2_acab
        Col 18: operadores_nec_L2_prod
    """
    refs = []
    refs_incompletas = []

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue

        ref_id = str(row[0]).strip()
        nome = str(row[1]).strip() if row[1] else ref_id
        bolos_por_caixa = _safe_int(row[2], default=1)
        familia = str(row[3]).strip().lower() if row[3] else "sem_familia"

        # -------------------------
        # Dados L1
        # -------------------------
        pode_L1 = _eh_sim(row[10])
        cadencia_L1_prod = _safe_float(row[11])
        ops_L1_acab = _safe_int(row[12])
        ops_L1_prod = _safe_int(row[13])

        if pode_L1:
            if cadencia_L1_prod is None:
                refs_incompletas.append(
                    (ref_id, "pode_L1=Sim mas sem cadência L1")
                )
            if ops_L1_prod is None:
                refs_incompletas.append(
                    (ref_id, "pode_L1=Sim mas sem operadores produção L1")
                )
            if ops_L1_acab is None:
                refs_incompletas.append(
                    (ref_id, "pode_L1=Sim mas sem operadores acabamento L1")
                )
        else:
            cadencia_L1_prod = 0
            ops_L1_prod = 0
            ops_L1_acab = 0

        # -------------------------
        # Dados L2
        # -------------------------
        pode_L2 = _eh_sim(row[14])
        cadencia_L2_prod = _safe_float(row[15])
        cadencia_L2_acab = _safe_float(row[16])
        ops_L2_acab = _safe_int(row[17])
        ops_L2_prod = _safe_int(row[18])

        if pode_L2:
            if cadencia_L2_prod is None:
                refs_incompletas.append(
                    (ref_id, "pode_L2=Sim mas sem cadência produção L2")
                )
            if cadencia_L2_acab is None:
                refs_incompletas.append(
                    (ref_id, "pode_L2=Sim mas sem cadência acabamento L2")
                )
            if ops_L2_prod is None:
                refs_incompletas.append(
                    (ref_id, "pode_L2=Sim mas sem operadores produção L2")
                )
            if ops_L2_acab is None:
                refs_incompletas.append(
                    (ref_id, "pode_L2=Sim mas sem operadores acabamento L2")
                )
        else:
            cadencia_L2_prod = 0
            cadencia_L2_acab = 0
            ops_L2_prod = 0
            ops_L2_acab = 0

        ref = {
            "id": ref_id,
            "nome": nome,
            "familia": familia,
            "bolos_por_caixa": bolos_por_caixa,

            # L0 não modelado nesta versão, mas guardado para futuro
            "lead_time_L0_dias": _safe_int(row[9], default=1),

            # L1
            "pode_L1": pode_L1,
            "cadencia_L1_prod": cadencia_L1_prod,
            "cadencia_L1_acab": cadencia_L1_prod,
            "ops_L1_prod": ops_L1_prod,
            "ops_L1_acab": ops_L1_acab,

            # L2
            "pode_L2": pode_L2,
            "cadencia_L2_prod": cadencia_L2_prod,
            "cadencia_L2_acab": cadencia_L2_acab,
            "ops_L2_prod": ops_L2_prod,
            "ops_L2_acab": ops_L2_acab,
        }

        refs.append(ref)

    return refs, refs_incompletas


# ============================================================
# LEITURA DA ABA 1_ESTRUTURA
# ============================================================

def _ler_aba_estrutura(ws):
    """
    Lê a aba 1_ESTRUTURA para extrair parâmetros gerais.

    Mapeia os valores conhecidos para chaves da instância e guarda o resto
    em `_extra` para inspeção posterior.
    """
    estrutura = {
        "n_dias": 5,
        "capacidade_linha_min": CAPACIDADE_LINHA_BRUTA_MIN,
        "tempo_limpeza_fim_dia_min": TEMPO_LIMPEZA_FIM_DIA_MIN,
        "operadores_limpeza": OPERADORES_LIMPEZA,
        "capacidade_L0_min": None,
        "lead_time_padrao_L0_L1L2_dias": None,
        "horario_inicio_L0": None,
        "horario_fim_L0": None,
        "n_fornos": None,
        "capacidade_fornos_min": None,
        "horario_inicio_L1_producao": None,
        "horario_fim_L1_producao": None,
        "capacidade_L1_producao_min": None,
        "horario_inicio_L1_acab": None,
        "horario_fim_L1_acab": None,
        "capacidade_L1_acab_min": None,
        "tempo_tunel_L1_min": None,
        "horario_inicio_L2_producao": None,
        "horario_fim_L2_producao": None,
        "capacidade_L2_producao_min": None,
        "horario_inicio_L2_acab": None,
        "horario_fim_L2_acab": None,
        "capacidade_L2_acab_min": None,
        "tempo_azoto_L2_min": None,
        "n_operadores_produtivos": None,
        "n_operadores": None,
        "operadores_rodam_L0_L1_L2": None,
        "_extra": {},
    }

    parametro_para_chave = {
        "número de dias úteis": ("n_dias", "int"),
        "numero de dias uteis": ("n_dias", "int"),
        "dias uteis": ("n_dias", "int"),
        "capacidade efetiva de l0 por dia (minutos)": ("capacidade_L0_min", "int"),
        "lead time padrão l0 → l1/l2 (dias)": ("lead_time_padrao_L0_L1L2_dias", "int"),
        "horário de início de l0": ("horario_inicio_L0", "time"),
        "horário de fim de l0": ("horario_fim_L0", "time"),
        "número de fornos disponíveis": ("n_fornos", "int"),
        "capacidade efetiva por forno por dia (minutos)": ("capacidade_fornos_min", "int"),
        "horário de início l1 produção": ("horario_inicio_L1_producao", "time"),
        "horário de fim l1 produção": ("horario_fim_L1_producao", "time"),
        "capacidade efetiva l1 produção (minutos)": ("capacidade_L1_producao_min", "int"),
        "horário de início l1 acabamento/embalamento": ("horario_inicio_L1_acab", "time"),
        "horário de fim l1 acabamento/embalamento": ("horario_fim_L1_acab", "time"),
        "capacidade efetiva l1 acabamento/embalamento (minutos)": ("capacidade_L1_acab_min", "int"),
        "tempo do túnel de arrefecimento l1 (minutos)": ("tempo_tunel_L1_min", "int"),
        "horário de início l2 produção": ("horario_inicio_L2_producao", "time"),
        "horário de fim l2 produção": ("horario_fim_L2_producao", "time"),
        "capacidade efetiva l2 produção (minutos)": ("capacidade_L2_producao_min", "int"),
        "horário de início l2 acabamento/embalamento": ("horario_inicio_L2_acab", "time"),
        "horário de fim l2 acabamento/embalamento": ("horario_fim_L2_acab", "time"),
        "capacidade efetiva l2 acabamento/embalamento (minutos)": ("capacidade_L2_acab_min", "int"),
        "tempo da câmara de azoto l2 (minutos)": ("tempo_azoto_L2_min", "int"),
        "número total de operadores produtivos": ("n_operadores_produtivos", "int"),
        "número total de operadores ": ("n_operadores", "int"),
        "número total de operadores": ("n_operadores", "int"),
        "os operadores rodam entre l0/l1/l2?": ("operadores_rodam_L0_L1_L2", "bool"),
    }

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue

        param = str(row[0]).strip().lower()
        valor = row[1]

        mapeamento = parametro_para_chave.get(param)
        if mapeamento is not None:
            chave, tipo = mapeamento
            if tipo == "int":
                estrutura[chave] = _safe_int(valor, default=estrutura[chave])
            elif tipo == "bool":
                estrutura[chave] = _eh_sim(valor)
            elif tipo == "time":
                estrutura[chave] = _safe_time(valor, default=estrutura[chave])
            else:
                estrutura[chave] = _safe_text(valor, default=estrutura[chave])
        elif valor is not None:
            estrutura["_extra"][param] = valor

    return estrutura


# ============================================================
# LEITURA DA ABA 4_OPERADORES
# ============================================================

def _ler_aba_operadores(ws):
    """
    Lê a aba 4_OPERADORES.

    Cabeçalho na linha 4:
        Col 0: operador_id
        Col 1: nome
        Col 2: no_pool
        Col 3-7: seg, ter, qua, qui, sex
    """
    operadores = []

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue

        no_pool = _eh_sim(row[2])

        if not no_pool:
            continue

        disponibilidade = [
            _safe_int(row[3 + i], default=0)
            for i in range(5)
        ]

        operadores.append({
            "id": str(row[0]).strip(),
            "nome": str(row[1]).strip() if row[1] else str(row[0]).strip(),
            "disponibilidade": disponibilidade,
        })

    return operadores


# ============================================================
# LEITURA DA ABA 5_COMPETENCIAS
# ============================================================

def _ler_aba_competencias(ws):
    """
    Lê a aba 5_COMPETENCIAS.

    Cabeçalho na linha 4:
        Col 0: operador_id
        Col 1: nome
        Col 2: L0
        Col 3: L1
        Col 4: L2
    """
    competencias = {}

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue

        op_id = str(row[0]).strip()

        competencias[op_id] = {
            "L0": str(row[2]).strip() if row[2] else "-",
            "L1": str(row[3]).strip() if row[3] else "-",
            "L2": str(row[4]).strip() if row[4] else "-",
        }

    return competencias


# ============================================================
# LEITURA DA ABA 3_SETUPS
# ============================================================

def _ler_aba_setups(ws, familias_todas):
    """
    Lê a aba 3_SETUPS.
    Onde estiver vazio, assume valor padrão.
    """
    matriz = {}

    header = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    familias_colunas = [
        str(c).strip().lower() if c else None
        for c in header[1:]
    ]

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue

        fam_de = str(row[0]).strip().lower()

        for j, fam_para in enumerate(familias_colunas):
            if fam_para is None:
                continue

            valor = row[j + 1] if j + 1 < len(row) else None

            if isinstance(valor, (int, float)):
                matriz[(fam_de, fam_para)] = float(valor)

    SETUP_PADRAO = 30
    SETUP_MESMA_FAMILIA = 5
    n_estimados = 0

    for f1 in familias_todas:
        for f2 in familias_todas:
            if (f1, f2) not in matriz:
                if f1 == f2:
                    matriz[(f1, f2)] = SETUP_MESMA_FAMILIA
                else:
                    matriz[(f1, f2)] = SETUP_PADRAO

                n_estimados += 1

    return matriz, n_estimados


# ============================================================
# PROCURA SINTÉTICA
# ============================================================

def _gerar_procura_sintetica(refs, n_dias, n_pedidos=15, seed=42):
    """
    Gera procura sintética para testes.

    Só escolhe referências que tenham pelo menos uma linha válida:
    - pode_L1=True e cadência L1 > 0
    ou
    - pode_L2=True e cadência L2 > 0

    Se a cadência for None, a referência não é considerada válida nessa linha.
    """
    random.seed(seed)

    refs_validas = [
        r for r in refs
        if (
            r["pode_L1"] and _valor_positivo(r["cadencia_L1_prod"])
        ) or (
            r["pode_L2"] and _valor_positivo(r["cadencia_L2_prod"])
        )
    ]

    if not refs_validas:
        print("⚠️ Nenhuma referência válida encontrada para gerar procura sintética.")
        return []

    procura = []

    for _ in range(n_pedidos):
        ref = random.choice(refs_validas)

        pedido = {
            "ref_id": ref["id"],
            "caixas_master": random.choice([100, 150, 200, 300, 500]),
            "data_entrega": random.randint(2, n_dias),
            "prioridade": random.choice(["Alta", "Media", "Baixa"]),
        }

        procura.append(pedido)

    return procura


# ============================================================
# CARREGADOR PRINCIPAL
# ============================================================

def carregar_instancia_real(
    caminho_excel="Inputs_Doceleia.xlsx",
    n_pedidos_sinteticos=15,
    seed=42
):
    """
    Lê o ficheiro Excel da Empresa X e devolve uma instância pronta a usar.
    """
    print(f"A carregar instância de {caminho_excel}...")

    wb = openpyxl.load_workbook(caminho_excel, data_only=True)

    estrutura = _ler_aba_estrutura(wb["1_ESTRUTURA"])
    refs, refs_incompletas = _ler_aba_referencias(wb["2_REFERENCIAS"])
    operadores = _ler_aba_operadores(wb["4_OPERADORES"])
    competencias = _ler_aba_competencias(wb["5_COMPETENCIAS"])

    familias = sorted(set(r["familia"] for r in refs))

    matriz_setups, n_setups_estimados = _ler_aba_setups(
        wb["3_SETUPS"],
        familias
    )

    procura = _gerar_procura_sintetica(
        refs,
        estrutura["n_dias"],
        n_pedidos=n_pedidos_sinteticos,
        seed=seed
    )

    instancia = {
        # Estrutura
        "n_dias": estrutura["n_dias"],
        "linhas_finais": ["L1", "L2"],
        "dias": [f"dia_{i + 1}" for i in range(estrutura["n_dias"])],
        "capacidade_linha_min": estrutura["capacidade_linha_min"],
        "tempo_limpeza_fim_dia_min": estrutura["tempo_limpeza_fim_dia_min"],
        "operadores_limpeza": estrutura["operadores_limpeza"],

        # Se a limpeza NÃO deve descontar ao tempo disponível, troca por:
        # "tempo_disponivel_linha_min": estrutura["capacidade_linha_min"],
        "tempo_disponivel_linha_min": (
            estrutura["capacidade_linha_min"]
            - estrutura["tempo_limpeza_fim_dia_min"]
        ),

        # Dados
        "refs": refs,
        "familias": familias,
        "matriz_setups": matriz_setups,
        "operadores": operadores,
        "competencias": competencias,
        "procura": procura,
        "estrutura": estrutura,

        # Metadados sobre qualidade
        "_meta": {
            "n_refs_total": len(refs),
            "n_refs_incompletas": len(refs_incompletas),
            "refs_incompletas": refs_incompletas,
            "n_familias": len(familias),
            "n_setups_estimados": n_setups_estimados,
            "fonte_procura": "sintética (Aba 6 vazia)",
        }
    }

    return instancia


# ============================================================
# RELATÓRIO DA INSTÂNCIA CARREGADA
# ============================================================

def imprimir_resumo_instancia(instancia):
    """Imprime um relatório da instância carregada."""
    meta = instancia["_meta"]

    print("=" * 70)
    print("INSTÂNCIA CARREGADA — RESUMO")
    print("=" * 70)

    print("\n📅 ESTRUTURA TEMPORAL")
    print(f"  Horizonte: {instancia['n_dias']} dias úteis")
    print(f"  Capacidade L1/L2 bruta: {instancia['capacidade_linha_min']} min/dia")
    print(
        f"  Limpeza fim de dia: {instancia['tempo_limpeza_fim_dia_min']} min "
        f"({instancia['operadores_limpeza']} operadores)"
    )
    print(
        f"  Disponível para produção+setups: "
        f"{instancia['tempo_disponivel_linha_min']} min/dia"
    )

    print("\n📦 REFERÊNCIAS")
    print(f"  Total: {meta['n_refs_total']}")

    n_pode_L1 = sum(1 for r in instancia["refs"] if r["pode_L1"])
    n_pode_L2 = sum(1 for r in instancia["refs"] if r["pode_L2"])
    n_so_L1 = sum(
        1 for r in instancia["refs"]
        if r["pode_L1"] and not r["pode_L2"]
    )
    n_so_L2 = sum(
        1 for r in instancia["refs"]
        if r["pode_L2"] and not r["pode_L1"]
    )
    n_ambas = sum(
        1 for r in instancia["refs"]
        if r["pode_L1"] and r["pode_L2"]
    )

    print(f"    Podem L1: {n_pode_L1} (só L1: {n_so_L1})")
    print(f"    Podem L2: {n_pode_L2} (só L2: {n_so_L2})")
    print(f"    Podem ambas: {n_ambas}")

    if meta["n_refs_incompletas"] > 0:
        print(f"  ⚠️  Refs incompletas: {meta['n_refs_incompletas']}")

        for ref_id, motivo in meta["refs_incompletas"][:15]:
            print(f"      - {ref_id}: {motivo}")

        if meta["n_refs_incompletas"] > 15:
            print("      ...")

    print("\n🏷️  FAMÍLIAS")
    print(f"  Total: {meta['n_familias']}")
    print(
        f"  Lista: {', '.join(instancia['familias'][:10])}"
        f"{'...' if len(instancia['familias']) > 10 else ''}"
    )

    print("\n⚙️  SETUPS")
    n_total_setups = len(instancia["matriz_setups"])
    n_reais = n_total_setups - meta["n_setups_estimados"]

    print(
        f"  Matriz: {meta['n_familias']}×{meta['n_familias']} "
        f"= {n_total_setups} valores"
    )
    print(f"  Preenchidos no Excel: {n_reais}")
    print(f"  ⚠️  Estimados (valor padrão): {meta['n_setups_estimados']}")

    print("\n👥 OPERADORES (no pool partilhado)")
    print(f"  Total: {len(instancia['operadores'])}")

    if instancia["operadores"]:
        disponib_por_dia = [
            sum(op["disponibilidade"][d] for op in instancia["operadores"])
            for d in range(instancia["n_dias"])
        ]

        dias_nomes = ["seg", "ter", "qua", "qui", "sex"][:instancia["n_dias"]]

        print(f"  Disponíveis por dia: {dict(zip(dias_nomes, disponib_por_dia))}")

    print("\n📋 PROCURA")
    print(f"  Pedidos: {len(instancia['procura'])} ({meta['fonte_procura']})")

    if instancia["procura"]:
        print("  Primeiros 5 pedidos:")

        for p in instancia["procura"][:5]:
            print(
                f"    {p['ref_id']}: {p['caixas_master']} caixas master, "
                f"entrega {p['data_entrega']}, prioridade {p['prioridade']}"
            )

    print("\n💡 EXEMPLO DE CÁLCULO DE TEMPO")

    ref_exemplo = next(
        (
            r for r in instancia["refs"]
            if _valor_positivo(r["cadencia_L1_prod"])
        ),
        None
    )

    if ref_exemplo:
        cadencia = ref_exemplo["cadencia_L1_prod"]
        bpc = ref_exemplo["bolos_por_caixa"]

        for n_caixas in [100, 300, 500]:
            t = calcular_tempo_producao(n_caixas, bpc, cadencia)

            print(
                f"  {ref_exemplo['id']}: {n_caixas} caixas × {bpc} bolos = "
                f"{n_caixas * bpc} bolos a {cadencia} b/h → {t:.0f} min"
            )

    print()


# ============================================================
# BLOCO DE TESTE
# ============================================================

if __name__ == "__main__":
    instancia = carregar_instancia_real(caminho_excel="Inputs_Doceleia.xlsx")
    imprimir_resumo_instancia(instancia)